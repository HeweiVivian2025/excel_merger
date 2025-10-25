import os
import uuid
import logging
import atexit
import shutil
import re
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import load_workbook
from openpyxl.styles import Protection

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 初始化Flask应用
template_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'templates'))
app = Flask(__name__, template_folder=template_dir)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB 限制
app.config['UPLOAD_FOLDER'] = '/tmp/uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx'}

# --------------------------- 从参考脚本整合的Excel处理函数 ---------------------------
def try_parse_date(date_str):
    """尝试解析日期字符串"""
    if not isinstance(date_str, str):
        return None
    date_str = date_str.strip()
    date_formats = [
        "%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", 
        "%Y年%m月%d日", "%Y/%-m/%-d", "%-m/%-d/%Y"
    ]
    for fmt in date_formats:
        try:
            parsed = datetime.strptime(date_str, fmt)
            return parsed.date()
        except ValueError:
            continue
    return None

def find_sheet_by_name(workbook, target_name):
    """通过名称模糊查找工作表（不区分大小写和空格）"""
    # 创建灵活的匹配模式
    target_pattern = re.compile(r'\s*'.join(re.escape(part) for part in target_name.split()), re.IGNORECASE)
    
    # 精确匹配（移除所有空格后比较）
    clean_target = re.sub(r'\s+', '', target_name).lower()
    for sheet_name in workbook.sheetnames:
        clean_name = re.sub(r'\s+', '', sheet_name).lower()
        if clean_name == clean_target:
            return workbook[sheet_name]
    
    # 模糊匹配
    for sheet_name in workbook.sheetnames:
        if target_pattern.search(sheet_name):
            return workbook[sheet_name]
    
    return None

def safe_cell_value(ws, row, col):
    """安全获取单元格值，处理合并单元格和数据清洗"""
    try:
        cell = ws.cell(row=row, column=col)
        
        # 处理合并单元格
        if ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # 返回合并区域左上角单元格的值
                    top_left_cell = ws.cell(
                        row=merged_range.min_row,
                        column=merged_range.min_col
                    )
                    value = top_left_cell.value
                    break
            else:
                value = cell.value
        else:
            value = cell.value
        
        # 数据清洗
        if value is None or value == "":
            return None
            
        if isinstance(value, str):
            value = value.strip()
            if value == "" or value.upper() in ["NONE", "#N/A", "NULL", "#VALUE!", "#REF!"]:
                return None
                
        # 尝试解析日期
        parsed_date = try_parse_date(str(value)) if isinstance(value, (str, datetime)) else None
        if parsed_date:
            return parsed_date
            
        return value
        
    except Exception as e:
        logger.error(f"读取单元格 R{row}C{col} 时出错: {e}")
        return None

def unmerge_cells_if_merged(ws, cell_coord):
    """解除单元格的合并状态"""
    try:
        cell = ws[cell_coord]
        for merged_range in list(ws.merged_cells.ranges):
            if cell.coordinate in merged_range:
                ws.unmerge_cells(str(merged_range))
                logger.info(f"已解除合并单元格: {merged_range}")
        return True
    except Exception as e:
        logger.error(f"解除合并单元格失败: {str(e)}")
        return False

# --------------------------- 原有功能函数 ---------------------------
def allowed_file(filename):
    """检查文件是否为允许的类型"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def cleanup_temp_files():
    """清理临时文件"""
    if os.path.exists(app.config['UPLOAD_FOLDER']):
        for filename in os.listdir(app.config['UPLOAD_FOLDER']):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    logger.info(f"清理临时文件: {filename}")
            except Exception as e:
                logger.error(f"清理临时文件失败: {str(e)}")

# 注册退出时清理函数
atexit.register(cleanup_temp_files)

# --------------------------- Flask路由 ---------------------------
@app.route('/')
def index():
    """渲染主页面"""
    try:
        return render_template('index.html')
    except Exception as e:
        logger.error(f"渲染模板失败: {str(e)}")
        return "服务器错误", 500

@app.route('/upload', methods=['POST'])
def upload_files():
    """处理文件上传和合并"""
    if 'template' not in request.files or 'sources' not in request.files:
        return jsonify({'error': '请求中缺少模板文件或源文件'}), 400
    
    template_file = request.files['template']
    source_files = request.files.getlist('sources')
    
    # 验证文件选择
    if template_file.filename == '':
        return jsonify({'error': '未选择模板文件'}), 400
    if len(source_files) == 0 or all(f.filename == '' for f in source_files):
        return jsonify({'error': '未选择源文件'}), 400
    
    # 验证文件类型
    if not allowed_file(template_file.filename):
        return jsonify({'error': '模板文件必须是.xlsx格式'}), 400
    for f in source_files:
        if f.filename != '' and not allowed_file(f.filename):
            return jsonify({'error': f'源文件 {f.filename} 必须是.xlsx格式'}), 400
    
    try:
        # 保存上传文件
        template_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(template_file.filename))
        template_file.save(template_path)
        
        source_paths = []
        for source_file in source_files:
            if source_file.filename:
                source_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(source_file.filename))
                source_file.save(source_path)
                source_paths.append(source_path)
        
        # 生成输出文件路径
        output_filename = f"merged_{uuid.uuid4().hex}.xlsx"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        
        # 复制模板文件
        shutil.copy(template_path, output_path)
        
        # 加载工作簿
        template_wb = load_workbook(output_path, read_only=False, data_only=True)
        template_ws = template_wb.active
        row_index = 2  # 从第2行开始写入数据
        
        # 处理每个源文件
        for source_path in source_paths:
            source_filename = os.path.basename(source_path)
            try:
                # 加载源文件
                source_wb = load_workbook(source_path, read_only=False, data_only=True)
                source_ws = source_wb.active
                
                # 使用参考脚本的safe_cell_value读取数据
                e_value = safe_cell_value(source_ws, 1, 5)  # E1 (第1行第5列)
                f_value = safe_cell_value(source_ws, 1, 6)  # F1 (第1行第6列)
                g_value = safe_cell_value(source_ws, 1, 7)  # G1 (第1行第7列)
                
                logger.info(f"读取源文件 {source_filename} 的值 - E1: {e_value}, F1: {f_value}, G1: {g_value}")
                
                # 处理模板文件中的目标单元格（解除合并）
                target_cells = [f'E{row_index}', f'F{row_index}', f'G{row_index}']
                for cell_coord in target_cells:
                    unmerge_cells_if_merged(template_ws, cell_coord)
                
                # 写入模板文件
                template_ws[f'E{row_index}'] = e_value if e_value is not None else ''
                template_ws[f'F{row_index}'] = f_value if f_value is not None else ''
                template_ws[f'G{row_index}'] = g_value if g_value is not None else ''
                
                row_index += 1
                logger.info(f"成功处理源文件: {source_filename}")
                
                # 关闭源文件
                source_wb.close()
                
            except Exception as e:
                logger.error(f"处理源文件 {source_filename} 失败: {str(e)}", exc_info=True)
                return jsonify({'error': f'处理源文件 {source_filename} 时出错: {str(e)}'}), 500
        
        # 确保工作表未受保护
        if template_ws.protection.sheet:
            template_ws.protection = Protection(sheet=False)
        
        # 保存合并后的文件
        template_wb.save(output_path)
        logger.info(f"成功保存合并文件: {output_path}")
        
        # 验证文件是否存在
        if not os.path.exists(output_path):
            raise Exception("合并文件保存后不存在")
            
        return send_file(
            output_path,
            as_attachment=True,
            download_name='merged_output.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"合并文件处理失败: {str(e)}", exc_info=True)
        return jsonify({'error': f'处理文件时出错: {str(e)}'}), 500

@app.errorhandler(500)
def internal_server_error(e):
    logger.error(f"服务器内部错误: {str(e)}")
    return jsonify({'error': '服务器内部错误，请稍后再试'}), 500

@app.errorhandler(400)
def bad_request(e):
    return jsonify({'error': '请求参数错误，请检查输入'}), 400

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': '请求的资源不存在'}), 404

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)